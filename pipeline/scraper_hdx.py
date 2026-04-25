"""
scraper_hdx.py — HDX (UN OCHA) / PSA tabular data scraper.

Downloads the official HDX Excel file containing the complete PSGC table
(Philippine Standard Geographic Code) from PSA/NAMRIA via UN OCHA HDX.

This is the AUTHORITATIVE PSGC source — it has all 1,642 ADM3 municipalities
across the Philippines, including all 53 Cebu municipalities.

HDX Source: https://data.humdata.org/dataset/cod-ab-phl
Direct URL: https://data.humdata.org/dataset/caf116df-f984-4deb-85ca-41b349d3f313/resource/e74fd350-3728-427f-8b4c-0589dc563c87/download/phl_adminboundaries_tabulardata.xlsx

Output:
  - data/raw/psgc_hdx.xlsx        (raw download)
  - data/normalized/psgc_cebu.csv (Cebu-only PSGC table)
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from common import OUTPUT_DIR, write_csv, print_banner

SOURCE_URL = "https://data.humdata.org/dataset/caf116df-f984-4deb-85ca-41b349d3f313/resource/e74fd350-3728-427f-8b4c-0589dc563c87/download/phl_adminboundaries_tabulardata.xlsx"
SOURCE_TAG = "HDX/PSA"


def fetch_hdx_xlsx() -> bytes | None:
    """Download HDX tabular Excel file."""
    try:
        import cloudscraper
    except ImportError:
        print("  cloudscraper not installed: pip install cloudscraper")
        return None

    print(f"  Downloading from: {SOURCE_URL}")
    try:
        scraper = cloudscraper.create_scraper()
        resp = scraper.get(SOURCE_URL, timeout=120)
        if resp.status_code == 200:
            print(f"  ✅ Downloaded {len(resp.content):,} bytes")
            return resp.content
        print(f"  HTTP {resp.status_code}")
        return None
    except Exception as e:
        print(f"  Error: {e}")
        return None


def parse_hdx_xlsx(xlsx_bytes: bytes) -> list[dict]:
    """
    Parse HDX Excel. Extracts all ADM3 rows and returns as list of dicts.
    Columns: ADM3_EN, ADM3_PCODE, ADM3_REF, ADM3ALT1EN, ADM2_EN, ADM2_PCODE, ADM1_EN, ADM1_PCODE, ...
    """
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed: pip install openpyxl")
        return []

    records = []
    import io
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    ws = wb['ADM3']

    # Get header
    header = [str(c) if c else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
    print(f"  HDX ADM3 columns: {header[:8]}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        record = dict(zip(header, row))
        records.append(record)

    wb.close()
    print(f"  Total ADM3 records: {len(records)}")
    return records


def filter_cebu(records: list[dict]) -> list[dict]:
    """Filter to Cebu Province (ADM2_PCODE = PH07022)."""
    cebu = [r for r in records if r.get("ADM2_PCODE") == "PH07022"]
    print(f"  Cebu ADM3 records: {len(cebu)}")
    return cebu


def psgc_from_hdx(hdx_code: str) -> str | None:
    """
    Convert HDX PSGC code (PH0702201) to standard PSGC format (072201).

    HDX format: PH + 2-digit region (07) + 3-digit province (022) + 2-digit municipality (01)
              = PH0702201
    Standard:  2-digit region (07) + 2-digit province (22) + 2-digit municipality (01)
             = 072201

    Conversion: strip PH, then take region(2) + prov[1:] + muni
                07 + 022 + 01 → 07 + 22 + 01 = 072201
                (i.e. skip the leading '0' of the 3-digit province code)
    """
    if not hdx_code or len(hdx_code) < 7:
        return None
    stripped = hdx_code.replace("PH", "").replace("ph", "")
    if len(stripped) == 7:
        region = stripped[:2]   # "07"
        prov   = stripped[2:5]  # "022"
        muni   = stripped[5:7]  # "01"
        # region + skip leading 0 of province + muni
        # "07" + "22" + "01" = "072201"
        return f"{region}{prov[1]}{prov[2]}{muni}"
    return None


def normalize_hdx_record(record: dict) -> dict:
    """Convert HDX record to standard PSGC format."""
    hdx_code = record.get("ADM3_PCODE", "")
    standard_psgc = psgc_from_hdx(hdx_code)
    return {
        "psgc_code": standard_psgc or "",
        "municipality": record.get("ADM3_EN", ""),
        "adm2_name": record.get("ADM2_EN", ""),
        "adm1_name": record.get("ADM1_EN", ""),
        "hdx_code": hdx_code,
        "alt_name": record.get("ADM3ALT1EN", ""),
        "source": SOURCE_TAG,
    }


def save_cebu_psgc(records: list[dict]) -> Path:
    """Save Cebu PSGC table to CSV."""
    normalized = [normalize_hdx_record(r) for r in records]
    out_path = OUTPUT_DIR / "normalized" / "psgc_cebu.csv"

    rows = []
    for r in sorted(normalized, key=lambda x: x["psgc_code"] or ""):
        rows.append({
            "psgc_code":    r["psgc_code"],
            "municipality": r["municipality"],
            "hdx_code":     r["hdx_code"],
            "alt_name":     r["alt_name"],
            "province":     r["adm2_name"],
            "region":       r["adm1_name"],
            "source":       r["source"],
        })

    write_csv(rows, out_path)
    return out_path


def main():
    print_banner("HDX/PSA PSGC Scraper")
    print("  Downloads official Philippine PSGC table from HDX/UN OCHA")
    print(f"  Source: {SOURCE_URL}")

    xlsx_bytes = fetch_hdx_xlsx()
    if not xlsx_bytes:
        print("\n  ❌ Failed to download HDX file.")
        sys.exit(1)

    # Save raw
    raw_path = OUTPUT_DIR / "raw" / "psgc_hdx.xlsx"
    raw_path.parent.mkdir(parents=True, exist_ok=True)
    with open(raw_path, "wb") as f:
        f.write(xlsx_bytes)
    print(f"  Saved raw → {raw_path}")

    records = parse_hdx_xlsx(xlsx_bytes)
    if not records:
        print("  ❌ No records parsed.")
        sys.exit(1)

    cebu = filter_cebu(records)
    if not cebu:
        print("  ❌ No Cebu records found.")
        sys.exit(1)

    out_path = save_cebu_psgc(cebu)
    print(f"\n✅ Done. {len(cebu)} Cebu municipalities saved to:")
    print(f"   {out_path}")

    # Print first few
    print("\nFirst 10 entries:")
    for r in sorted(cebu, key=lambda x: x.get("ADM3_PCODE", ""))[:10]:
        hdx = r.get("ADM3_PCODE", "")
        psgc = psgc_from_hdx(hdx)
        print(f"  {psgc} | {r.get('ADM3_EN', '')}")


if __name__ == "__main__":
    main()
